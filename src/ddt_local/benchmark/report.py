"""Benchmark report generation: CSV, Excel, terminal ranking."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ddt_local.benchmark.runner import BenchmarkSummary, DocumentRunResult


def _sanitize_excel_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    if text and text[0] in {"=", "+", "-", "@"}:
        return "'" + text
    return text


def write_csv_report(summary: BenchmarkSummary, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "run_name",
        "pipeline",
        "ocr_model",
        "struct_model",
        "vision_model",
        "document",
        "repetition",
        "success",
        "weighted_score",
        "header_accuracy",
        "lines_found",
        "lines_missing",
        "lines_invented",
        "wrong_quantities",
        "total_latency_seconds",
        "retry_count",
        "error_message",
    ]
    with output_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for r in summary.results:
            writer.writerow(
                {
                    "run_name": r.run_name,
                    "pipeline": r.pipeline,
                    "ocr_model": r.ocr_model or "",
                    "struct_model": r.struct_model or "",
                    "vision_model": r.vision_model or "",
                    "document": r.document_filename,
                    "repetition": r.repetition,
                    "success": r.success,
                    "weighted_score": round(r.score.weighted_score, 4),
                    "header_accuracy": round(r.score.header_accuracy, 4),
                    "lines_found": r.score.lines_found,
                    "lines_missing": r.score.lines_missing,
                    "lines_invented": r.score.lines_invented,
                    "wrong_quantities": r.score.wrong_quantities,
                    "total_latency_seconds": round(r.total_latency_seconds, 3),
                    "retry_count": r.retry_count,
                    "error_message": r.error_message or "",
                }
            )
    return output_path


def write_excel_report(summary: BenchmarkSummary, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Per-run sheets
    by_run: dict[str, list[DocumentRunResult]] = {}
    for r in summary.results:
        by_run.setdefault(r.run_name, []).append(r)

    first = True
    for run_name, rows in by_run.items():
        if first:
            ws = wb.active
            ws.title = _safe_sheet_name(run_name)
            first = False
        else:
            ws = wb.create_sheet(_safe_sheet_name(run_name))
        headers = [
            "document",
            "repetition",
            "weighted_score",
            "header_accuracy",
            "lines_found",
            "lines_missing",
            "lines_invented",
            "latency_s",
            "success",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append(
                [
                    _sanitize_excel_value(r.document_filename),
                    r.repetition,
                    round(r.score.weighted_score, 4),
                    round(r.score.header_accuracy, 4),
                    r.score.lines_found,
                    r.score.lines_missing,
                    r.score.lines_invented,
                    round(r.total_latency_seconds, 3),
                    r.success,
                ]
            )
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        _autosize(ws)

    # Comparative sheet
    comparative = wb.create_sheet("Comparativo")
    run_names = list(by_run.keys())
    means = summary.mean_scores_by_run()
    metrics = [
        ("weighted_score_mean", means),
        ("docs", {n: len(by_run[n]) for n in run_names}),
        (
            "avg_latency_s",
            {
                n: (
                    sum(r.total_latency_seconds for r in by_run[n]) / len(by_run[n])
                    if by_run[n]
                    else 0
                )
                for n in run_names
            },
        ),
        (
            "avg_lines_missing",
            {
                n: (
                    sum(r.score.lines_missing for r in by_run[n]) / len(by_run[n])
                    if by_run[n]
                    else 0
                )
                for n in run_names
            },
        ),
    ]
    comparative.append(["metric", *run_names])
    for cell in comparative[1]:
        cell.font = Font(bold=True)
    for metric_name, values in metrics:
        comparative.append([metric_name, *[round(values.get(n, 0), 4) for n in run_names]])
    comparative.auto_filter.ref = comparative.dimensions
    comparative.freeze_panes = "B2"
    _autosize(comparative)

    # Errors sheet
    errors = wb.create_sheet("Errori")
    errors.append(["document", "run_name", "field", "expected", "predicted", "error_type", "weight"])
    for cell in errors[1]:
        cell.font = Font(bold=True)
    for r in summary.results:
        for err in r.score.error_details:
            errors.append(
                [
                    _sanitize_excel_value(r.document_filename),
                    _sanitize_excel_value(r.run_name),
                    _sanitize_excel_value(err.field),
                    _sanitize_excel_value(err.expected),
                    _sanitize_excel_value(err.predicted),
                    _sanitize_excel_value(err.error_type),
                    err.weight,
                ]
            )
    errors.auto_filter.ref = errors.dimensions
    errors.freeze_panes = "A2"
    _autosize(errors)

    wb.save(output_path)
    return output_path


def format_ranking_text(summary: BenchmarkSummary) -> str:
    lines = ["=== Benchmark ranking (mean weighted_score) ==="]
    if not summary.ranking:
        lines.append("(no results)")
        return "\n".join(lines)
    for i, (name, score) in enumerate(summary.ranking, start=1):
        lines.append(f"{i}. {name}: {score:.4f}")
    return "\n".join(lines)


def write_reports(summary: BenchmarkSummary, output_dir: Path) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = write_csv_report(summary, output_dir / f"benchmark_{stamp}.csv")
    xlsx_path = write_excel_report(summary, output_dir / f"benchmark_{stamp}.xlsx")
    return csv_path, xlsx_path


def _safe_sheet_name(name: str) -> str:
    cleaned = "".join(c if c.isalnum() or c in " _-" else "_" for c in name)
    return cleaned[:31] or "run"


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max_len + 2, 48)
