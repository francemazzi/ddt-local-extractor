"""Excel export for the production DDT archive."""

from __future__ import annotations

import os
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ddt_local.database import Database


def sanitize_excel_value(value: object) -> object:
    """Return a safe Excel cell value, neutralising formula-like user content."""
    if value is None:
        return ""
    if isinstance(value, (int, float, Decimal, date)):
        return value
    text = str(value)
    if text.lstrip()[:1] in {"=", "+", "-", "@"}:
        return "'" + text
    return text


def write_production_excel(database: Database, output_path: Path) -> Path:
    """Build the four-sheet user archive and atomically replace ``output_path``."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    _write_ddt_sheet(workbook.active, database.production_headers())
    _write_lines_sheet(workbook.create_sheet("Righe"), database.production_lines())
    _write_errors_sheet(workbook.create_sheet("Errori"), database.production_errors())
    _write_review_sheet(workbook.create_sheet("Da verificare"), database.production_review_items())

    temporary_path: Path | None = None
    try:
        with NamedTemporaryFile(
            mode="wb", prefix=f".{output_path.stem}.", suffix=".xlsx", dir=output_path.parent, delete=False
        ) as temporary:
            temporary_path = Path(temporary.name)
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
        temporary_path = None
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)
    return output_path


def _write_ddt_sheet(ws, rows: Iterable[Any]) -> None:
    ws.title = "DDT"
    headers = [
        "File sorgente",
        "Numero DDT",
        "Data DDT",
        "Riferimento ordine",
        "Causale",
        "Fornitore",
        "P. IVA fornitore",
        "Destinatario",
        "P. IVA destinatario",
        "Numero colli",
        "Peso lordo",
        "Peso netto",
        "Vettore",
        "Destinazione",
        "Quality score",
        "Da verificare",
        "SHA-256",
        "Elaborato il",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                sanitize_excel_value(row["source_filename"]),
                sanitize_excel_value(row["numero_ddt"]),
                _as_excel_date(row["data_ddt"]),
                sanitize_excel_value(row["riferimento_ordine"]),
                sanitize_excel_value(row["causale_trasporto"]),
                sanitize_excel_value(row["fornitore_ragione_sociale"]),
                sanitize_excel_value(row["fornitore_partita_iva"]),
                sanitize_excel_value(row["destinatario_ragione_sociale"]),
                sanitize_excel_value(row["destinatario_partita_iva"]),
                _as_excel_number(row["numero_colli"]),
                _as_excel_number(row["peso_lordo"]),
                _as_excel_number(row["peso_netto"]),
                sanitize_excel_value(row["vettore"]),
                sanitize_excel_value(row["destinazione"]),
                row["quality_score"],
                bool(row["requires_review"]),
                sanitize_excel_value(row["sha256"]),
                sanitize_excel_value(row["finished_at"]),
            ]
        )
    _finish_sheet(ws, date_columns={3})


def _write_lines_sheet(ws, rows: Iterable[Any]) -> None:
    headers = [
        "File sorgente",
        "Numero DDT",
        "Data DDT",
        "Riga",
        "Codice",
        "Descrizione",
        "Quantità",
        "Unità di misura",
        "Lotto",
        "Matricola",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                sanitize_excel_value(row["source_filename"]),
                sanitize_excel_value(row["numero_ddt"]),
                _as_excel_date(row["data_ddt"]),
                row["line_index"],
                sanitize_excel_value(row["codice"]),
                sanitize_excel_value(row["descrizione"]),
                _as_excel_number(row["quantita"]),
                sanitize_excel_value(row["unita_misura"]),
                sanitize_excel_value(row["lotto"]),
                sanitize_excel_value(row["matricola"]),
            ]
        )
    _finish_sheet(ws, date_columns={3})


def _write_errors_sheet(ws, rows: Iterable[Any]) -> None:
    headers = [
        "File sorgente",
        "SHA-256",
        "Stato",
        "Errore documento",
        "Campo",
        "Tipo",
        "Gravità",
        "Dettaglio",
        "Elaborato il",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                sanitize_excel_value(row["original_filename"]),
                sanitize_excel_value(row["sha256"]),
                sanitize_excel_value(row["status"]),
                sanitize_excel_value(row["error_message"]),
                sanitize_excel_value(row["field_path"]),
                sanitize_excel_value(row["issue_type"]),
                sanitize_excel_value(row["severity"]),
                sanitize_excel_value(row["description"]),
                sanitize_excel_value(row["finished_at"]),
            ]
        )
    _finish_sheet(ws)


def _write_review_sheet(ws, rows: Iterable[Any]) -> None:
    headers = ["File sorgente", "Numero DDT", "Data DDT", "Quality score", "Problemi"]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                sanitize_excel_value(row["source_filename"]),
                sanitize_excel_value(row["numero_ddt"]),
                _as_excel_date(row["data_ddt"]),
                row["quality_score"],
                sanitize_excel_value(row["issues"]),
            ]
        )
    _finish_sheet(ws, date_columns={3})


def _finish_sheet(ws, *, date_columns: set[int] | None = None) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 48)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width
    for column_number in date_columns or set():
        for cell in list(ws.columns)[column_number - 1][1:]:
            if isinstance(cell.value, date):
                cell.number_format = "DD/MM/YYYY"


def _as_excel_date(value: object) -> date | object:
    if value in (None, ""):
        return ""
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value))
    except ValueError:
        return sanitize_excel_value(value)


def _as_excel_number(value: object) -> float | int | object:
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)):
        return value
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return sanitize_excel_value(value)
    if decimal_value == decimal_value.to_integral_value():
        return int(decimal_value)
    return float(decimal_value)
