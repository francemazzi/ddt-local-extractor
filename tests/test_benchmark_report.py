"""Tests for benchmark report."""

from __future__ import annotations

from ddt_local.benchmark.report import (
    format_ranking_text,
    write_csv_report,
    write_excel_report,
)
from ddt_local.benchmark.runner import BenchmarkSummary, DocumentRunResult
from ddt_local.benchmark.scoring import ErrorDetail, ScoreResult


def _summary() -> BenchmarkSummary:
    score = ScoreResult(
        weighted_score=0.8,
        header_accuracy=0.9,
        lines_found=2,
        lines_missing=0,
        lines_invented=0,
        error_details=[
            ErrorDetail(
                document="a.pdf",
                field="=cmd",
                expected="x",
                predicted="+y",
                error_type="formatting",
                weight=0.1,
            )
        ],
    )
    result = DocumentRunResult(
        run_name="ocr_qwen4b",
        pipeline="ocr_struct",
        ocr_model="glm-ocr:latest",
        struct_model="qwen3.5:4b",
        vision_model=None,
        document_filename="a.pdf",
        repetition=1,
        score=score,
        success=True,
        total_latency_seconds=1.5,
    )
    summary = BenchmarkSummary(results=[result], ranking=[("ocr_qwen4b", 0.8)])
    return summary


def test_write_csv(tmp_path):
    path = write_csv_report(_summary(), tmp_path / "out.csv")
    text = path.read_text(encoding="utf-8")
    assert "ocr_qwen4b" in text
    assert "weighted_score" in text


def test_write_excel_sheets_and_injection(tmp_path):
    path = write_excel_report(_summary(), tmp_path / "out.xlsx")
    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert "ocr_qwen4b" in wb.sheetnames or any("ocr" in s.lower() for s in wb.sheetnames)
    assert "Comparativo" in wb.sheetnames
    assert "Errori" in wb.sheetnames
    errors = wb["Errori"]
    # formula injection neutralized with leading apostrophe
    field_val = str(errors.cell(2, 3).value)
    assert field_val.startswith("'") or not field_val.startswith("=")


def test_ranking_text():
    text = format_ranking_text(_summary())
    assert "ocr_qwen4b" in text
    assert "0.8000" in text
