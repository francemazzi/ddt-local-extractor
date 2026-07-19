"""Tests for production Excel export and result persistence."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from ddt_local.database import Database
from ddt_local.excel import sanitize_excel_value, write_production_excel
from ddt_local.models import (
    DocumentoDDT,
    DatiDocumento,
    ExecutionMetadata,
    ExtractionArtifacts,
    ExtractionMethod,
    ExtractionResult,
    RigaDDT,
    Soggetto,
    SourceDocument,
)


@pytest.fixture
def database(tmp_path: Path) -> Database:
    db = Database(tmp_path / "data" / "ddt.sqlite3")
    db.initialize()
    return db


def _success_result(*, requires_review: bool = True) -> ExtractionResult:
    document = DocumentoDDT(
        source_filename="=formula.pdf",
        fornitore=Soggetto(ragione_sociale="+Fornitore", partita_iva="IT123"),
        destinatario=Soggetto(ragione_sociale="Destinatario", partita_iva="IT456"),
        documento=DatiDocumento(
            numero_ddt="@DDT-42",
            data_ddt=date(2026, 7, 18),
            numero_colli="12",
            peso_lordo="101.5",
            peso_netto="100.25",
        ),
        articoli=[
            RigaDDT(
                numero_riga=1,
                codice="=CODE",
                descrizione="-descrizione",
                quantita="4.5",
                unita_misura="PZ",
            )
        ],
        quality_score=0.72 if requires_review else 0.95,
        campi_da_verificare=["documento.numero_ddt"] if requires_review else [],
    )
    return ExtractionResult(
        document=document,
        metadata=ExecutionMetadata(
            pipeline="ocr_struct",
            page_count=1,
            extraction_method=ExtractionMethod.NATIVE_TEXT,
            total_duration_seconds=1.5,
        ),
        artifacts=ExtractionArtifacts(
            ocr_text_by_page=["OCR content for persistence"],
            table_markdown="| codice | quantità |\n| --- | --- |\n| ART-1 | 4.5 |",
            raw_json_response='{"safe": true}',
        ),
        success=True,
    )


def _source(tmp_path: Path, sha256: str = "a" * 64) -> SourceDocument:
    source = tmp_path / "=formula.pdf"
    source.write_bytes(b"PDF")
    return SourceDocument(
        path=source,
        filename=source.name,
        sha256=sha256,
        size_bytes=source.stat().st_size,
        page_count=1,
    )


def test_excel_has_required_sheets_dates_numbers_and_safe_cells(database: Database, tmp_path: Path):
    database.persist_extraction_result(_source(tmp_path), _success_result())
    output = write_production_excel(database, tmp_path / "output" / "DDT_estratti.xlsx")

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["DDT", "Righe", "Errori", "Da verificare"]

    ddt = workbook["DDT"]
    assert ddt["A2"].value == "'=formula.pdf"
    assert ddt["B2"].value == "'@DDT-42"
    assert ddt["C2"].value == datetime(2026, 7, 18)
    assert ddt["C2"].number_format == "DD/MM/YYYY"
    assert ddt["J2"].value == 12
    assert ddt["K2"].value == 101.5

    lines = workbook["Righe"]
    assert lines["E2"].value == "'=CODE"
    assert lines["F2"].value == "'-descrizione"
    assert lines["G2"].value == 4.5
    assert workbook["Da verificare"].max_row == 2


def test_excel_write_is_atomic_when_save_fails(database: Database, tmp_path: Path):
    database.persist_extraction_result(_source(tmp_path), _success_result())
    output = tmp_path / "output" / "DDT_estratti.xlsx"
    output.parent.mkdir()
    output.write_bytes(b"previous workbook")

    with patch("ddt_local.excel.Workbook.save", side_effect=RuntimeError("disk full")):
        with pytest.raises(RuntimeError, match="disk full"):
            write_production_excel(database, output)

    assert output.read_bytes() == b"previous workbook"
    assert list(output.parent.glob(".DDT_estratti.*.xlsx")) == []


def test_persistence_rolls_back_everything_when_line_insert_fails(
    database: Database, tmp_path: Path
):
    source = _source(tmp_path)
    with patch.object(database, "insert_ddt_lines", side_effect=RuntimeError("insert failed")):
        with pytest.raises(RuntimeError, match="insert failed"):
            database.persist_extraction_result(source, _success_result())

    assert database.document_exists_by_hash(source.sha256) is False
    assert database.count_production_documents() == 0


def test_persistence_archives_ocr_pages_lines_and_validation_issues(
    database: Database, tmp_path: Path
):
    database.persist_extraction_result(_source(tmp_path), _success_result())

    with database.connect() as conn:
        assert conn.execute("SELECT COUNT(*) FROM ddt_headers").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM ddt_lines").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM ocr_pages").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM validation_issues").fetchone()[0] == 1


@pytest.mark.parametrize("value", ["=1+1", "+1", "-1", "@formula"])
def test_sanitize_excel_value_neutralizes_formula_injection(value: str):
    assert sanitize_excel_value(value) == "'" + value
